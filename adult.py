################################
# READ ME
# Execute with :
#     python adult.py mnode mdepth
# Example :
#     python adult.py 300 20
################################
import pandas as pd
import sys
# Load the max_node and max_depth from argument
# max_node and max_depth is for decision tree
mnode = int(sys.argv[1])
mdepth = int(sys.argv[2])
print('max node : ', mnode, '  max depth : ', mdepth)
# Load origin dataset by pandas from UCI, load dataset-'adult'
origin = pd.read_csv('https://archive.ics.uci.edu/ml/machine-learning-databases/adult/adult.data', names=['age','workclass','fnlwgt','education','education-num','marital-status','occupation','relationship','race','sex','capital-gain','capital-loss','hours-per-week','native-country','salary'])
# Drop the unnecessary(I thought) column
origin = origin.drop('fnlwgt', axis=1)
# Define the encode tag
encode_tag = ['workclass', 'education', 'marital-status', 'occupation', 'relationship', 'race', 'sex', 'native-country']

# Decision tree cannot train by string attributes, need to transform
from sklearn.preprocessing import OneHotEncoder as ohe
import numpy as np
# Initial
enc = ohe(categories='auto')
# Transform attributes column by column
for i in encode_tag:
    # Load single column and transform it
    t = pd.DataFrame(data=(enc.fit_transform(origin[i].to_numpy().reshape(-1,1)).toarray()))
    # Replace the origin column by transformed column
    origin = pd.concat([origin.drop(i, axis=1), t], axis=1)
#print(origin)

# Split origin dataset to training and testing dataset
from sklearn.model_selection import train_test_split as tts
train, test = tts(origin, test_size=0.2)
# Split origin dataset to value and label
trainX = train.drop('salary', axis=1).sort_index()
trainY = pd.DataFrame(data=(enc.fit_transform(train['salary'].to_numpy().reshape(-1,1)).toarray()))
# Split origin dataset to value and label
testX = test.drop('salary', axis=1).sort_index()
testY = pd.DataFrame(data=(enc.fit_transform(test['salary'].to_numpy().reshape(-1,1)).toarray()))

# Training part
from sklearn import tree
# Initial classifier with input arguments
dt = tree.DecisionTreeClassifier(max_leaf_nodes=mnode, max_depth=mdepth)
# Train
dt = dt.fit(trainX, trainY)
# Predict training data
train_predict = dt.predict(trainX)
# Predict testing data
test_predict = pd.DataFrame(data=dt.predict(testX))

# Save the decision tree
with open('adult.dot', 'w') as f:
    f = tree.export_graphviz(dt, out_file=f)

# Check the accuracy
from sklearn import metrics
train_acc = metrics.accuracy_score(trainY, train_predict)
test_acc = metrics.accuracy_score(testY, test_predict)
print('Accuracy on training data is %f' % train_acc)
print('Accuracy on testing data is %f' % test_acc)

# Write the label and predict result to excel
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows as dtr
filename = 'adult_result.xls'
wb = Workbook()
compare = pd.concat([testY[0], test_predict.rename(columns={0:'1'})['1']], axis=1)
compare = compare.rename(columns={0:'class','1':'predict'})
sheet1 = wb.create_sheet('test', 0)
for x in dtr(compare):
    sheet1.append(x)
wb.save(filename)
